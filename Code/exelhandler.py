import openpyxl,pprint



house1 = openpyxl.load_workbook('2011_August_04_to_18.xlsx')
house1_array = []
sheet = house1.get_sheet_by_name('Trend')
for row in range(10, 298):
	
	state = sheet['M' + str(row)].value
	house1_array.append(state)
print(house1_array)
print (len(house1_array))

house2 = openpyxl.load_workbook('2011_June_03_to_17.xlsx')
house2_array = []
sheet = house2.get_sheet_by_name('Trend')
for row in range(10, 298):
	
	state = sheet['M' + str(row)].value
	house2_array.append(state)
print(house2_array)
print (len(house2_array))

house3 = openpyxl.load_workbook('2012_July_16_to_26.xlsx')
house3_array = []
sheet = house3.get_sheet_by_name('Trend')
for row in range(10, 298):
	
	state = sheet['AU' + str(row)].value
	house3_array.append(state)
print(house3_array)
print (len(house3_array))

house4 = openpyxl.load_workbook('2012_June_01_to_15.xlsx')
house4_array = []
sheet = house4.get_sheet_by_name('Trend')
for row in range(10, 298):
	
	state = sheet['M' + str(row)].value
	house4_array.append(state)
print(house4_array)
print (len(house4_array))

house5 = openpyxl.load_workbook('Data House10_1minutes_01-06-2012_30-06-2012.xlsx')
house5_array = []
sheet = house5.get_sheet_by_name('Trend')
for row in range(2, 1440, 5):
	
	state = sheet['H' + str(row)].value
	house5_array.append(state)
print(house5_array)
print (len(house5_array))

office1 = openpyxl.load_workbook('Office.xlsx')
office1_array = []
sheet = office1.get_sheet_by_name('Sheet1')
for row in range(2, 290):
	
	state = sheet['B' + str(row)].value
	office1_array.append(state)
print(office1_array)
print (len(office1_array))

office2 = openpyxl.load_workbook('office2.xlsx')
office2_array = []
sheet = office2.get_sheet_by_name('Sheet1')
for row in range(2, 8640, 30):
	
	state = sheet['A' + str(row)].value
	office2_array.append(state)
print(office2_array)
print (len(office2_array))

bar = openpyxl.load_workbook('Data Commercial1_1minute_02-07-2014_07-07-2014.xlsx')
bar_array = []
sheet = bar.get_sheet_by_name('Trend')
for row in range(10, 1450, 5):
	
	state = sheet['AR' + str(row)].value
	bar_array.append(state)
print(bar_array)

print (len(bar_array))

total_consumption = []
for i in range (0,288,1):
	
	total_consumption.append(bar_array[i] + office2_array[i] + house5_array[i]+ house1_array[i] + house2_array[i]+house3_array[i]+ house4_array[i]+office1_array[i]) 

for i in range(len(total_consumption)):
	
	print(total_consumption[i])
print(total_consumption)	
	
	
wind = openpyxl.load_workbook('wind_generation.xlsx')
wind_array = []
sheet = wind.get_sheet_by_name('Sheet1')
for row in range(1, 145):
	
	state = sheet['A' + str(row)].value
	wind_array.append(state)
	
print (len(wind_array))

wind_array_new = []

for i in range (0,144):
	
	if (i==143):
		wind_array_new.append (wind_array[i])
	else :
		wind_array_new.append (wind_array[i])
		wind_array_new.append((wind_array[i] + wind_array[i+1])/2)


solar = openpyxl.load_workbook('Weather data Sao Paulo_USP-ENERQ_5minutes_01-12-2013_10-12-2013.xlsx')
solar_array = []
sheet = solar.get_sheet_by_name('Trend')
for row in range(3, 290):
	
	state = sheet['I' + str(row)].value
	solar_array.append(state)
	
print (solar_array)

total_gen = []
for i in range (0,287):
 total_gen.append (solar_array[i] + wind_array_new[i])

print (total_gen)



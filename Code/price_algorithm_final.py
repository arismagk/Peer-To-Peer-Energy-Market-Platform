from mpl_toolkits.mplot3d import Axes3D
from matplotlib import cm
from matplotlib.ticker import LinearLocator, FormatStrFormatter
import numpy as np
import scipy.io
import matplotlib.pyplot as plt
import math
import openpyxl,pprint


house5 = openpyxl.load_workbook('Data House10_1minutes_01-06-2012_30-06-2012.xlsx')
house5_array = []
sheet = house5.get_sheet_by_name('Trend')
for row in range(2, 1440, 5):
	
	state = sheet['H' + str(row)].value
	house5_array.append(state)
print(house5_array)
print (len(house5_array))


office2 = openpyxl.load_workbook('office2.xlsx')
office2_array = []
sheet = office2.get_sheet_by_name('Sheet1')
for row in range(2, 8640, 30):
	
	state = sheet['A' + str(row)].value
	office2_array.append(state)
print(office2_array)
print (len(office2_array))

bar = openpyxl.load_workbook('Data Commercial1_1minute_02-07-2014_07-07-2014.xlsx')
bar_array = []
sheet = bar.get_sheet_by_name('Trend')
for row in range(10, 1450, 5):
	
	state = sheet['AR' + str(row)].value
	bar_array.append(state)
print(bar_array)

print (len(bar_array))

total_consumption = []
for i in range (0,288,1):
	
	total_consumption.append(bar_array[i] + office2_array[i] + house5_array[i]) 


wind = openpyxl.load_workbook('wind_generation.xlsx')
wind_array = []
sheet = wind.get_sheet_by_name('Sheet1')
for row in range(1, 145):
	
	state = sheet['A' + str(row)].value
	wind_array.append(state)
	
print (len(wind_array))

wind_array_new = []

for i in range (0,144):
	
	if (i==143):
		wind_array_new.append (wind_array[i])
	else :
		wind_array_new.append (wind_array[i])
		wind_array_new.append((wind_array[i] + wind_array[i+1])/2)


solar = openpyxl.load_workbook('Weather data Sao Paulo_USP-ENERQ_5minutes_01-12-2013_10-12-2013.xlsx')
solar_array = []
sheet = solar.get_sheet_by_name('Trend')
for row in range(3, 290):
	
	state = sheet['I' + str(row)].value
	solar_array.append(state)
	
print (solar_array)

total_gen = []
for i in range (0,287):
 total_gen.append (solar_array[i] + wind_array_new[i])

print (total_gen)

def distance(x1,y1,x2,y2):
    return math.sqrt((x1-x2)**2+(y1-y2)**2)

def get_price(production,consumption,buy_price):
    length=len(buy_price)-1
    while production>length or consumption>length:
        production/=2.0
        consumption/=2.0
    

    low_x=int(math.floor(production))
    low_y=int(math.floor(consumption))

    high_x=int(math.ceil(production))
    high_y=int(math.ceil(consumption))

    sum=0.0
    total_weight=0.0
    if low_x<steps and low_x>=0 and low_y<steps and low_y>=0 :
        sum+=buy_price[low_x][low_y]*1.0/(0.01+distance(low_x,low_y,production,consumption))
        total_weight+=1.0/(0.01+distance(low_x,low_y,production,consumption))
    if high_x<steps and high_x>=0 and low_y<steps and low_y>=0 :
        sum+=buy_price[high_x][low_y]*1.0/(0.01+distance(high_x,low_y,production,consumption))
        total_weight+=1.0/(0.01+distance(high_x,low_y,production,consumption))
    if low_x<steps and low_x>=0 and high_y<steps and high_y>=0 :
        sum+=buy_price[low_x][high_y]*1.0/(0.01+distance(low_x,high_y,production,consumption))
        total_weight+=1.0/(0.01+distance(low_x,high_y,production,consumption))
    if high_x<steps and high_x>=0 and high_y<steps and high_y>=0 :
        sum+=buy_price[high_x][high_y]*1.0/(0.01+distance(high_x,high_y,production,consumption))
        total_weight+=1.0/(0.01+distance(high_x,high_y,production,consumption))
    return round(sum/total_weight,2)

steps=500
debug=False

if (steps%2==0) :
    print('steps have to be odd')


internal_consumption=[0]*steps
internal_production=[0]*steps
buy_price=[[]]*steps
for i in range(steps):
    buy_price[i]=[0]*steps

deh_sells_electricity = 28.69	
deh_buys_electricity = 12.31

middle_price=(deh_sells_electricity+deh_buys_electricity)/2

buy_price[0][steps-1]=deh_sells_electricity

buy_price[steps-1][0]=deh_buys_electricity

interpolation_offset=middle_price
interpolation_top=deh_sells_electricity-interpolation_offset
interpolation_bottom=deh_buys_electricity-interpolation_offset


print("maximum price after offset:"+str(interpolation_top))
print("minimum price after offset:"+str(interpolation_bottom))

#interpolation diagonally copy pasting the main diagonal
for i in range(0,steps,1) :
    diagonal_size=i+1
    if debug:
        print("------ diagonal size:"+str(diagonal_size)+" ---------")
    if diagonal_size is 1:
        buy_price[0][0]=interpolation_offset
        continue

    distance_to_interpolation_edge=(diagonal_size-1.0)/2.0

    interpolation_factor_at_edge=interpolation_top**(1.0/3.0)
    interpolation_factor=interpolation_factor_at_edge/distance_to_interpolation_edge

    if debug:
        print("int factor"+str(interpolation_factor))

    for k in range(diagonal_size):

        write_y=i-k
        write_x=k
        if debug:
            print("trying to write to y:"+str(write_y)+" x:"+str(write_x))

        if write_y>=steps or write_y<0 or write_x>=steps or write_x<0 :
            print("oob")
        else:
            to_write=round((interpolation_factor*(k-(diagonal_size-1)/2.0))**3.0+interpolation_offset,2)
            buy_price[write_x][write_y]=to_write
            if debug:
                print("wrote :"+str(to_write))

for i in range(steps-2,-1,-1) :
    diagonal_size=i+1
    if debug:
        print("------ diagonal size:"+str(diagonal_size)+" ---------")
    if diagonal_size is 1:
        buy_price[steps-1][steps-1]=interpolation_offset
        continue

    distance_to_interpolation_edge=(diagonal_size-1)/2.0

    interpolation_factor_at_edge=interpolation_top**(1.0/3.0)
    interpolation_factor=interpolation_factor_at_edge/distance_to_interpolation_edge
    if debug:
        print("int factor"+str(interpolation_factor))

    for k in range(diagonal_size):

        write_y=steps-1-k
        write_x=steps-1-(i-k)
        if debug:
            print("tried to write to y:"+str(write_y)+" x:"+str(write_x))

        if write_y>=steps or write_y<0 or write_x>=steps or write_x<0 :
           
            print("oob")
        else:
            to_write=round((interpolation_factor*(k-(diagonal_size-1)/2.0))**3.0+interpolation_offset,2)
            buy_price[write_x][write_y]=to_write
            if debug:
                print("wrote :"+str(to_write))

#visualization
matrix=np.matrix(buy_price)

X = np.arange(1, 1+steps,1)
Y = np.arange(1, 1+steps,1)
X,Y=np.meshgrid(X,Y)
fig = plt.figure()
ax = fig.gca(projection='3d')

m = cm.ScalarMappable(cmap=cm.jet)
m.set_array(matrix)

surf =ax.plot_surface(X,Y,matrix, cmap=cm.coolwarm,
                       linewidth=0.1, antialiased=False)


# Add a color bar which maps values to colors.
fig.colorbar(surf, shrink=0.5, aspect=5)

plt.show()
if False:
    scipy.io.savemat('c:/tmp/arrdata.mat', mdict={'matrix': matrix})

price_array =[]
for i in range (0,287):
	print(str(get_price(total_consumption[i],total_gen[i],buy_price)))
	price_array.append(get_price(total_consumption[i],total_gen[i],buy_price))
	

print('OK')
print (len(price_array))
